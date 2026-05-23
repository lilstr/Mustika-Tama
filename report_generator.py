"""
Module untuk generate laporan PDF dan Excel
"""
from datetime import datetime
from io import BytesIO
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.pdfgen import canvas


def get_db_connection():
    return mysql.connector.connect(
        host='localhost',
        user='root',
        password='',
        database='db_panti_mustika'
    )


def format_currency(value):
    """Format angka ke format mata uang Indonesia"""
    return f"Rp {value:,.0f}".replace(",", ".")


def generate_penghuni_excel(period_type='all', year=None, month=None):
    """
    Generate laporan penghuni dalam format Excel
    period_type: 'all', 'yearly', 'monthly'
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    
    # Query data penghuni aktif
    query = """
    SELECT p.* FROM penghuni p
    LEFT JOIN mantan_penghuni mp ON mp.penghuni_id = p.id
    WHERE mp.penghuni_id IS NULL
    """

    if period_type == 'yearly' and year:
        query += f" AND YEAR(tanggal_masuk) = {year}"
    elif period_type == 'monthly' and year and month:
        query += f" AND YEAR(tanggal_masuk) = {year} AND MONTH(tanggal_masuk) = {month}"

    query += " ORDER BY nama_anak"
    
    cur.execute(query)
    data = cur.fetchall()
    cur.close()
    conn.close()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Penghuni"
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 25
    
    # Header styling
    header_fill = PatternFill(start_color="FF059669", end_color="FF059669", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    title = ws['A1']
    title.value = "LAPORAN DATA PENGHUNI PANTI ASUHAN MUSTIKA TAMA"
    title.font = Font(bold=True, size=12)
    title.alignment = Alignment(horizontal="center", vertical="center")
    
    # Period info
    ws.merge_cells('A2:H2')
    if period_type == 'yearly':
        period_text = f"Laporan Tahunan - Tahun {year}"
    elif period_type == 'monthly':
        month_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                       'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
        period_text = f"Laporan Bulanan - {month_names[month]} {year}"
    else:
        period_text = f"Laporan Semua Data - {datetime.now().strftime('%d/%m/%Y')}"
    
    period_cell = ws['A2']
    period_cell.value = period_text
    period_cell.font = Font(italic=True, size=10)
    period_cell.alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ['No', 'Nama Anak', 'NIK', 'Jenis Kelamin', 'Usia', 'Asal', 'Pendidikan', 'Tanggal Masuk']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, row_data in enumerate(data, 5):
        ws.cell(row=row_idx, column=1).value = row_idx - 4
        ws.cell(row=row_idx, column=2).value = row_data.get('nama_anak', '')
        ws.cell(row=row_idx, column=3).value = row_data.get('nik', '')
        ws.cell(row=row_idx, column=4).value = 'Laki-laki' if row_data.get('jenis_kelamin') == 'L' else 'Perempuan'
        ws.cell(row=row_idx, column=5).value = row_data.get('usia', '')
        ws.cell(row=row_idx, column=6).value = row_data.get('asal', '')
        ws.cell(row=row_idx, column=7).value = row_data.get('pendidikan', '')
        ws.cell(row=row_idx, column=8).value = row_data.get('tanggal_masuk', '').strftime('%d/%m/%Y') if row_data.get('tanggal_masuk') else ''
        
        # Apply borders to all cells
        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).border = border
            if col == 1 or col == 4:
                ws.cell(row=row_idx, column=col).alignment = center_alignment
    
    # Summary
    summary_row = len(data) + 6
    ws.merge_cells(f'A{summary_row}:B{summary_row}')
    summary_cell = ws[f'A{summary_row}']
    summary_cell.value = "Total Penghuni"
    summary_cell.font = Font(bold=True)
    
    ws[f'C{summary_row}'].value = len(data)
    ws[f'C{summary_row}'].font = Font(bold=True)
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Return BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def generate_donasi_excel(period_type='all', year=None, month=None):
    """
    Generate laporan donasi dalam format Excel
    period_type: 'all', 'yearly', 'monthly'
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    
    # Query data donasi
    query = "SELECT * FROM donasi WHERE status = 'paid'"
    
    if period_type == 'yearly' and year:
        query += f" AND YEAR(created_at) = {year}"
    elif period_type == 'monthly' and year and month:
        query += f" AND YEAR(created_at) = {year} AND MONTH(created_at) = {month}"
    
    query += " ORDER BY created_at DESC"
    
    cur.execute(query)
    data = cur.fetchall()
    cur.close()
    conn.close()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Donasi"
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20
    
    # Header styling
    header_fill = PatternFill(start_color="FF06b6d4", end_color="FF06b6d4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:G1')
    title = ws['A1']
    title.value = "LAPORAN DONASI PANTI ASUHAN MUSTIKA TAMA"
    title.font = Font(bold=True, size=12)
    title.alignment = Alignment(horizontal="center", vertical="center")
    
    # Period info
    ws.merge_cells('A2:G2')
    if period_type == 'yearly':
        period_text = f"Laporan Tahunan - Tahun {year}"
    elif period_type == 'monthly':
        month_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                       'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
        period_text = f"Laporan Bulanan - {month_names[month]} {year}"
    else:
        period_text = f"Laporan Semua Data - {datetime.now().strftime('%d/%m/%Y')}"
    
    period_cell = ws['A2']
    period_cell.value = period_text
    period_cell.font = Font(italic=True, size=10)
    period_cell.alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ['No', 'Tanggal Donasi', 'Nama Donatur', 'Kota', 'No. HP', 'Jumlah Donasi', 'Doa']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    
    total_donasi = 0
    
    for row_idx, row_data in enumerate(data, 5):
        ws.cell(row=row_idx, column=1).value = row_idx - 4
        ws.cell(row=row_idx, column=2).value = row_data.get('created_at', '').strftime('%d/%m/%Y') if row_data.get('created_at') else ''
        
        # Nama donatur (anonim atau nama)
        nama = row_data.get('nama_donatur', '')
        if row_data.get('anonim'):
            nama = "Donatur Anonim"
        ws.cell(row=row_idx, column=3).value = nama
        
        ws.cell(row=row_idx, column=4).value = row_data.get('kota', '')
        ws.cell(row=row_idx, column=5).value = row_data.get('no_hp', '')
        
        jumlah = row_data.get('jumlah', 0)
        total_donasi += jumlah
        ws.cell(row=row_idx, column=6).value = jumlah
        ws.cell(row=row_idx, column=6).number_format = '#,##0'
        
        ws.cell(row=row_idx, column=7).value = row_data.get('doa', '')
        
        # Apply borders
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = border
            if col == 1 or col == 2:
                ws.cell(row=row_idx, column=col).alignment = center_alignment
            elif col == 6:
                ws.cell(row=row_idx, column=col).alignment = right_alignment
    
    # Summary
    summary_row = len(data) + 6
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    summary_cell = ws[f'A{summary_row}']
    summary_cell.value = f"Total Donasi ({len(data)} Donatur)"
    summary_cell.font = Font(bold=True, size=11)
    summary_cell.border = border
    
    total_cell = ws[f'F{summary_row}']
    total_cell.value = total_donasi
    total_cell.font = Font(bold=True, size=11)
    total_cell.number_format = '#,##0'
    total_cell.border = border
    total_cell.fill = PatternFill(start_color="FFe0f2fe", end_color="FFe0f2fe", fill_type="solid")
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Return BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def generate_penghuni_pdf(period_type='all', year=None, month=None):
    """
    Generate laporan penghuni dalam format PDF
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    
    # Query data penghuni aktif
    query = """
    SELECT p.* FROM penghuni p
    LEFT JOIN mantan_penghuni mp ON mp.penghuni_id = p.id
    WHERE mp.penghuni_id IS NULL
    """

    if period_type == 'yearly' and year:
        query += f" AND YEAR(tanggal_masuk) = {year}"
    elif period_type == 'monthly' and year and month:
        query += f" AND YEAR(tanggal_masuk) = {year} AND MONTH(tanggal_masuk) = {month}"

    query += " ORDER BY nama_anak"
    
    cur.execute(query)
    data = cur.fetchall()
    cur.close()
    conn.close()
    
    # Create PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#059669'),
        spaceAfter=6,
        alignment=1
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=12,
        alignment=1
    )
    
    # Title
    story.append(Paragraph("LAPORAN DATA PENGHUNI", title_style))
    story.append(Paragraph("PANTI ASUHAN MUSTIKA TAMA", title_style))
    
    # Period info
    if period_type == 'yearly':
        period_text = f"Laporan Tahunan - Tahun {year}"
    elif period_type == 'monthly':
        month_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                       'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
        period_text = f"Laporan Bulanan - {month_names[month]} {year}"
    else:
        period_text = f"Laporan Semua Data - {datetime.now().strftime('%d/%m/%Y')}"
    
    story.append(Paragraph(period_text, normal_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Table data
    table_data = [['No', 'Nama Anak', 'NIK', 'Jenis Kelamin', 'Usia', 'Asal', 'Pendidikan']]
    
    for idx, row in enumerate(data, 1):
        table_data.append([
            str(idx),
            row.get('nama_anak', '')[:20],
            row.get('nik', '')[:15],
            'L' if row.get('jenis_kelamin') == 'L' else 'P',
            str(row.get('usia', '')),
            row.get('asal', '')[:15],
            row.get('pendidikan', '')[:10]
        ])
    
    # Create table
    table = Table(table_data, colWidths=[0.5*inch, 1.5*inch, 1.2*inch, 1*inch, 0.6*inch, 1*inch, 1.1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph(f"Total Penghuni: <b>{len(data)}</b>", normal_style))
    
    doc.build(story)
    output.seek(0)
    
    return output


def generate_donasi_pdf(period_type='all', year=None, month=None):
    """
    Generate laporan donasi dalam format PDF
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    
    # Query data donasi
    query = "SELECT * FROM donasi WHERE status = 'paid'"
    
    if period_type == 'yearly' and year:
        query += f" AND YEAR(created_at) = {year}"
    elif period_type == 'monthly' and year and month:
        query += f" AND YEAR(created_at) = {year} AND MONTH(created_at) = {month}"
    
    query += " ORDER BY created_at DESC"
    
    cur.execute(query)
    data = cur.fetchall()
    cur.close()
    conn.close()
    
    # Create PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#06b6d4'),
        spaceAfter=6,
        alignment=1
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=12,
        alignment=1
    )
    
    # Title
    story.append(Paragraph("LAPORAN DONASI", title_style))
    story.append(Paragraph("PANTI ASUHAN MUSTIKA TAMA", title_style))
    
    # Period info
    if period_type == 'yearly':
        period_text = f"Laporan Tahunan - Tahun {year}"
    elif period_type == 'monthly':
        month_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                       'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
        period_text = f"Laporan Bulanan - {month_names[month]} {year}"
    else:
        period_text = f"Laporan Semua Data - {datetime.now().strftime('%d/%m/%Y')}"
    
    story.append(Paragraph(period_text, normal_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Table data
    table_data = [['No', 'Tanggal', 'Nama', 'Kota', 'Jumlah']]
    
    total = 0
    for idx, row in enumerate(data, 1):
        nama = row.get('nama_donatur', '')
        if row.get('anonim'):
            nama = "Anonim"
        
        jumlah = row.get('jumlah', 0)
        total += jumlah
        
        tgl = row.get('created_at', '').strftime('%d/%m/%Y') if row.get('created_at') else ''
        
        table_data.append([
            str(idx),
            tgl,
            nama[:25],
            row.get('kota', '')[:15],
            f"Rp {jumlah:,}"
        ])
    
    # Create table
    table = Table(table_data, colWidths=[0.5*inch, 1*inch, 2*inch, 1.2*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#06b6d4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph(f"Total Donasi: <b>Rp {total:,}</b> ({len(data)} Donatur)", normal_style))
    
    doc.build(story)
    output.seek(0)
    
    return output
